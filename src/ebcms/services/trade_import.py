import csv
import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any

from pydantic import ValidationError
from sqlalchemy.orm import Session

from ebcms.config import get_settings
from ebcms.core.enums import TradeStatus
from ebcms.models import AuditLog, Client, Product, Trade, TradeImportBatch, TradeImportReject
from ebcms.schemas import TradeCreate, TradeImportRejectedRow, TradeImportSummary
from ebcms.services.validation import validate_trade_record

REQUIRED_COLUMNS = {
    "trade_id",
    "client_id",
    "product_id",
    "quantity",
    "price",
    "currency",
    "exchange",
    "trade_side",
    "trade_date",
}


class TradeImportError(ValueError):
    """Raised when an uploaded file cannot be parsed as a trade import."""


@dataclass(frozen=True)
class RawTradeRow:
    row_number: int
    data: dict[str, Any]


def import_trade_file(
    *,
    filename: str,
    content: bytes,
    imported_by: str,
    db: Session,
) -> TradeImportSummary:
    rows, source_type = parse_trade_file(filename, content)
    batch = TradeImportBatch(
        filename=filename,
        source_type=source_type,
        status="PROCESSING",
        total_rows=len(rows),
        imported_by=imported_by,
    )
    db.add(batch)
    db.flush()

    imported_trade_ids: list[str] = []
    rejected_records: list[TradeImportRejectedRow] = []
    seen_trade_ids: set[str] = set()

    for raw_row in rows:
        trade_payload, parse_issues = _parse_trade_payload(raw_row.data)
        trade_id = _clean(raw_row.data.get("trade_id"))

        if trade_payload is None:
            _record_rejection(
                batch=batch,
                raw_row=raw_row,
                trade_id=trade_id,
                issues=parse_issues,
                rejected_records=rejected_records,
                db=db,
            )
            continue

        validation_issues = _validate_import_payload(trade_payload, seen_trade_ids, db)
        if validation_issues:
            _record_rejection(
                batch=batch,
                raw_row=raw_row,
                trade_id=trade_payload.trade_id,
                issues=validation_issues,
                rejected_records=rejected_records,
                db=db,
            )
            continue

        trade = Trade(
            trade_id=trade_payload.trade_id,
            client_id=trade_payload.client_id,
            product_id=trade_payload.product_id,
            quantity=trade_payload.quantity,
            price=trade_payload.price,
            currency=trade_payload.currency.upper(),
            exchange=trade_payload.exchange.upper(),
            trade_side=str(trade_payload.trade_side).upper(),
            trade_date=trade_payload.trade_date,
            import_id=batch.import_id,
            status=TradeStatus.VALIDATED.value,
        )
        db.add(trade)
        db.add(
            AuditLog(
                entity_type="TRADE",
                entity_id=trade.trade_id,
                action="IMPORT",
                new_value=f"import_id={batch.import_id}; status={trade.status}",
                user_id=imported_by,
            )
        )
        imported_trade_ids.append(trade.trade_id)
        seen_trade_ids.add(trade.trade_id.upper())

    batch.accepted_rows = len(imported_trade_ids)
    batch.rejected_rows = len(rejected_records)
    batch.status = "COMPLETED_WITH_REJECTIONS" if rejected_records else "COMPLETED"
    db.commit()
    db.refresh(batch)

    return TradeImportSummary(
        import_id=batch.import_id,
        filename=batch.filename,
        source_type=batch.source_type,
        status=batch.status,
        total_rows=batch.total_rows,
        accepted_rows=batch.accepted_rows,
        rejected_rows=batch.rejected_rows,
        imported_trade_ids=imported_trade_ids,
        rejected_records=rejected_records,
    )


def parse_trade_file(filename: str, content: bytes) -> tuple[list[RawTradeRow], str]:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "csv":
        return _parse_csv(content), "CSV"
    if extension in {"xlsx", "xlsm"}:
        return _parse_xlsx(content), "XLSX"
    raise TradeImportError("Unsupported file type. Upload a .csv, .xlsx, or .xlsm trade file.")


def _parse_csv(content: bytes) -> list[RawTradeRow]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise TradeImportError("CSV file must include a header row.")

    headers = [_normalize_header(header) for header in reader.fieldnames]
    _validate_required_columns(headers)

    rows: list[RawTradeRow] = []
    for index, row in enumerate(reader, start=2):
        normalized = {_normalize_header(key): _normalize_cell(value) for key, value in row.items()}
        if any(value not in {"", None} for value in normalized.values()):
            rows.append(RawTradeRow(row_number=index, data=normalized))
    return rows


def _parse_xlsx(content: bytes) -> list[RawTradeRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise TradeImportError("Excel imports require the openpyxl package.") from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        raise TradeImportError("Excel file must include a header row.")

    headers = [_normalize_header(header) for header in header_cells]
    _validate_required_columns(headers)

    rows: list[RawTradeRow] = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        normalized = {
            header: _normalize_cell(value)
            for header, value in zip(headers, values, strict=False)
            if header
        }
        if any(value not in {"", None} for value in normalized.values()):
            rows.append(RawTradeRow(row_number=index, data=normalized))
    workbook.close()
    return rows


def _validate_required_columns(headers: list[str]) -> None:
    missing = sorted(REQUIRED_COLUMNS.difference(set(headers)))
    if missing:
        raise TradeImportError(f"Missing required columns: {', '.join(missing)}.")


def _parse_trade_payload(raw_data: dict[str, Any]) -> tuple[TradeCreate | None, list[str]]:
    try:
        payload = TradeCreate.model_validate(
            {column: raw_data.get(column) for column in REQUIRED_COLUMNS}
        )
    except ValidationError as exc:
        issues = [
            f"{'.'.join(str(part) for part in error['loc'])}: {error['msg']}"
            for error in exc.errors()
        ]
        return None, issues
    return payload, []


def _validate_import_payload(
    payload: TradeCreate,
    seen_trade_ids: set[str],
    db: Session,
) -> list[str]:
    client = db.get(Client, payload.client_id)
    product = db.get(Product, payload.product_id)
    duplicate_trade_id = payload.trade_id.upper() in seen_trade_ids or db.get(Trade, payload.trade_id)

    validation = validate_trade_record(
        trade_id=payload.trade_id,
        client_exists=bool(client and client.is_active),
        product_exists=bool(product and product.is_active),
        duplicate_trade_id=bool(duplicate_trade_id),
        quantity=payload.quantity,
        price=payload.price,
        currency=payload.currency,
        trade_side=str(payload.trade_side),
        trade_date=payload.trade_date,
        allowed_currencies=get_settings().currency_set,
    )
    return validation.issues


def _record_rejection(
    *,
    batch: TradeImportBatch,
    raw_row: RawTradeRow,
    trade_id: str | None,
    issues: list[str],
    rejected_records: list[TradeImportRejectedRow],
    db: Session,
) -> None:
    reason = "; ".join(issues) if issues else "Unknown import error."
    raw_payload = _json_ready(raw_row.data)
    db.add(
        TradeImportReject(
            import_id=batch.import_id,
            row_number=raw_row.row_number,
            trade_id=trade_id,
            reason=reason,
            raw_payload=json.dumps(raw_payload, sort_keys=True),
        )
    )
    rejected_records.append(
        TradeImportRejectedRow(
            row_number=raw_row.row_number,
            trade_id=trade_id,
            reason=reason,
            raw_payload=raw_payload,
        )
    )


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _normalize_cell(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if value is None:
        return ""
    return str(value).strip()


def _clean(value: Any) -> str | None:
    cleaned = str(value or "").strip()
    return cleaned or None


def _json_ready(data: dict[str, Any]) -> dict[str, object]:
    return {key: _normalize_cell(value) for key, value in data.items()}
